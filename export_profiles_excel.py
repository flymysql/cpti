import json
import os
from glob import glob

base = '/Users/jimmy/Downloads/cpti'
profiles_dir = os.path.join(base, 'profiles')
json_files = sorted(glob(os.path.join(profiles_dir, '*.json')))
rows = []
for path in json_files:
    if os.path.basename(path) == 'index.js':
        continue
    with open(path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    rows.append({
        'id': data.get('id', ''),
        'abbr': data.get('abbr', ''),
        'name': data.get('name', ''),
        'accent': data.get('accent', ''),
        'tags': ' / '.join(data.get('tags', [])),
        'description': data.get('description', ''),
        'longDescription': data.get('longDescription', ''),
        'note': data.get('note', ''),
        'needDescription': data.get('needDescription', ''),
        'needLongDescription': data.get('needLongDescription', ''),
        'needNote': data.get('needNote', ''),
        'detailEssay': data.get('detailEssay', ''),
    })

rows = sorted(rows, key=lambda r: r['id'])

xlsx_path = os.path.join(base, '恋爱人格_26.xlsx')

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'profiles'
    headers = list(rows[0].keys()) if rows else []
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22
    wb.save(xlsx_path)
    print(xlsx_path)
except Exception:
    import csv
    csv_path = os.path.join(base, '恋爱人格_26.csv')
    headers = list(rows[0].keys()) if rows else []
    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
    print(csv_path)
